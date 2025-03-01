from flask import Flask, render_template, request, send_file, after_this_request, session
import os
import time
import pandas as pd
from datetime import timedelta
from threading import Thread
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import requests
import re

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
TURNSTILE_SECRET_KEY = os.getenv("TURNSTILE_SECRET_KEY")
app.secret_key = '@Your_secret_Key'

# 設置安全的會話配置
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=60)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = True  # 僅在 HTTPS 下啟用

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# 檔案清理函數
def delete_old_files(folder, age_in_seconds):
    now = time.time()
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        if os.path.isfile(file_path) and (now - os.path.getmtime(file_path)) > age_in_seconds:
            os.remove(file_path)
            print(f"Deleted: {file_path}")

def cleanup_task():
    while True:
        delete_old_files(UPLOAD_FOLDER, 60)
        delete_old_files(RESULT_FOLDER, 60)
        time.sleep(60)

# 時間轉秒數的輔助函數
def time_to_seconds(time_str):
    if not time_str or not isinstance(time_str, str):
        return 0
    days = hours = minutes = seconds = 0
    if "天" in time_str:
        days = int(re.search(r"(\d+)天", time_str).group(1)) if re.search(r"(\d+)天", time_str) else 0
    if "小時" in time_str or "時" in time_str:
        hours = int(re.search(r"(\d+)[時小時]", time_str).group(1)) if re.search(r"(\d+)[時小時]", time_str) else 0
    if "分" in time_str:
        minutes = int(re.search(r"(\d+)分", time_str).group(1)) if re.search(r"(\d+)分", time_str) else 0
    if "秒" in time_str:
        seconds = int(re.search(r"(\d+)秒", time_str).group(1)) if re.search(r"(\d+)秒", time_str) else 0
    return days * 86400 + hours * 3600 + minutes * 60 + seconds

# 秒數轉時分的輔助函數
def seconds_to_time(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return f"{hours:02d}時{minutes:02d}分"

@app.route('/')
def homepage():
    return render_template('home.html')

@app.route('/legacy')
def legacy():
    show_captcha = 'captcha_verified' not in session or not session['captcha_verified']
    return render_template('legacy.html', show_captcha=show_captcha)

@app.route('/legacy/upload', methods=['POST'])
def legacy_upload():
    # CAPTCHA 驗證
    if 'captcha_verified' not in session or not session['captcha_verified']:
        captcha_token = request.form.get('cf-turnstile-response')
        if not captcha_token:
            return {"error": "CAPTCHA token is missing"}, 400
        verification_url = "https://challenges.cloudflare.com/turnstile/v0/siteverify"
        payload = {"secret": TURNSTILE_SECRET_KEY, "response": captcha_token, "remoteip": request.remote_addr}
        response = requests.post(verification_url, data=payload)
        if not response.json().get("success"):
            return {"error": "CAPTCHA validation failed"}, 403
        session['captcha_verified'] = True
        session.permanent = True

    if 'file' not in request.files or not request.files['file'].filename:
        return {"error": "請上傳檔案"}, 400

    file = request.files['file']
    filename = secure_filename(file.filename)
    if not (filename.lower().endswith('.xlsx') or filename.lower().endswith('.csv')):
        return {"error": "檔案格式不支援，請上傳 .xlsx 或 .csv 檔案"}, 400

    try:
        input_path = os.path.join(UPLOAD_FOLDER, filename)
        output_filename = f"{os.path.splitext(filename)[0]}-count.xlsx"
        output_path = os.path.join(RESULT_FOLDER, output_filename)
        file.save(input_path)

        if filename.startswith('OnlineInfo') and filename.lower().endswith('.xlsx'):
            wb = load_workbook(input_path)
            ws = wb.active
            ws["E2"] = "秒數"
            ws["G2"] = "時分"
            for i in range(3, ws.max_row + 1):
                time_str = ws[f"C{i}"].value
                seconds = time_to_seconds(time_str)
                ws[f"E{i}"] = seconds
                ws[f"G{i}"] = seconds_to_time(seconds)
            wb.save(output_path)

        elif filename.startswith('ScoreReport') and filename.lower().endswith('.xlsx'):
            wb = load_workbook(input_path)
            ws = wb.active
            ws["G2"] = "MyET秒數"
            for i in range(3, ws.max_row + 1):
                time_str = ws[f"D{i}"].value
                ws[f"G{i}"] = time_to_seconds(time_str)
            wb.save(output_path)

        else:  # EasyTest CSV
            df = pd.read_csv(input_path, encoding='utf-8', encoding_errors='replace')
            if '總時數' in df.columns:
                df['總時數'] = df['總時數'].astype(str)
                df['秒數'] = df['總時數'].apply(time_to_seconds)
            if '登入次數' in df.columns:
                df = df.drop(columns=['登入次數'])
            df.to_excel(output_path, index=False)

        @after_this_request
        def remove_files(response):
            for path in [input_path, output_path]:
                if os.path.exists(path):
                    os.remove(path)
            return response

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        if os.path.exists(input_path):
            os.remove(input_path)
        return {"error": str(e)}, 500

@app.route('/new')
def new():
    show_captcha = 'captcha_verified' not in session or not session['captcha_verified']
    return render_template('new.html', show_captcha=show_captcha)

@app.route('/new/upload', methods=['POST'])
def upload_file():
    if 'captcha_verified' not in session or not session['captcha_verified']:
        captcha_token = request.form.get('cf-turnstile-response')
        if not captcha_token:
            return {"error": "CAPTCHA token is missing"}, 400
        verification_url = "https://challenges.cloudflare.com/turnstile/v0/siteverify"
        payload = {
            "secret": TURNSTILE_SECRET_KEY,
            "response": captcha_token,
            "remoteip": request.remote_addr
        }
        response = requests.post(verification_url, data=payload)
        if not response.json().get("success"):
            return {"error": "CAPTCHA validation failed"}, 403
        session['captcha_verified'] = True
        session.permanent = True

    if 'history_file' not in request.files or 'online_info_file' not in request.files:
        return "請上傳 history*.csv 和 OnlineInfo_*.xlsx 檔案", 400

    history_file = request.files['history_file']
    online_info_file = request.files['online_info_file']
    if not history_file.filename or not online_info_file.filename:
        return "沒有檔案被選取", 400

    if not history_file.filename.lower().endswith('.csv'):
        return "History 檔案必須是 .csv 格式", 400
    if not online_info_file.filename.lower().endswith('.xlsx'):
        return "OnlineInfo 檔案必須是 .xlsx 格式", 400

    try:
        history_filename = secure_filename(history_file.filename)
        online_info_filename = secure_filename(online_info_file.filename)
        history_path = os.path.join(UPLOAD_FOLDER, history_filename)
        online_info_path = os.path.join(UPLOAD_FOLDER, online_info_filename)
        history_file.save(history_path)
        online_info_file.save(online_info_path)
        result_path = os.path.join(RESULT_FOLDER, 'result.xlsx')

        history_df = pd.read_csv(history_path, encoding='utf-8', encoding_errors='replace')
        required_cols = ["使用者帳號", "總時數"]
        if not all(col in history_df.columns for col in required_cols):
            return f"History 檔案缺少必要的欄位: {required_cols}", 400
        history_df = history_df[required_cols]
        history_df['總時數'] = history_df['總時數'].astype(str)

        if online_info_filename.endswith("Stud_List.xlsx"):
            stud_df = pd.read_excel(online_info_path, sheet_name=0, skiprows=4, usecols=[0, 1, 2, 3],
                                    names=["班級", "學號", "姓名", "修別"])
            stud_df = stud_df.dropna(subset=["學號"])
            merged_df = pd.merge(stud_df, history_df, left_on="學號", right_on="使用者帳號", how="left")
            result_df = merged_df[["班級", "學號", "姓名", "修別", "總時數"]]
            result_df.columns = ["班級", "學號", "姓名", "修別", "EasyTest總時數"]
            result_df['EasyTest秒數'] = result_df['EasyTest總時數'].apply(time_to_seconds)
            result_df['成績'] = result_df['EasyTest秒數'].apply(lambda x: 10 if x >= 72000 else (x / 72000 * 10 if x > 0 else 0))
            result_df.to_excel(result_path, index=False)

        elif online_info_filename.startswith("ScoreReport_"):
            online_df = pd.read_excel(online_info_path, sheet_name=0, skiprows=1,
                                      usecols=["帳號", "名字", "上線時間", "平均分數", "次"])
            merged_df = pd.merge(online_df, history_df, left_on="帳號", right_on="使用者帳號", how="left")
            result_df = merged_df[["帳號", "名字", "上線時間", "總時數"]]
            result_df.columns = ["帳號", "名字", "MyET上線時間", "EasyTest總時數"]
            result_df['MyET秒數'] = result_df['MyET上線時間'].apply(time_to_seconds)
            result_df['EasyTest秒數'] = result_df['EasyTest總時數'].apply(time_to_seconds)
            result_df.to_excel(result_path, index=False)

        elif online_info_filename.startswith("OnlineInfo"):
            online_df = pd.read_excel(online_info_path, sheet_name=0, skiprows=1,
                                      usecols=["帳號", "姓名", "總上線時間", "登入學習次數"])
            merged_df = pd.merge(online_df, history_df, left_on="帳號", right_on="使用者帳號", how="left")
            result_df = merged_df[["帳號", "姓名", "總上線時間", "登入學習次數", "總時數"]]
            result_df.columns = ["帳號", "姓名", "MyET總上線時間", "登入學習次數", "EasyTest總時數"]
            result_df['MyET秒數'] = result_df['MyET總上線時間'].apply(time_to_seconds)
            result_df['MyET時分'] = result_df['MyET秒數'].apply(seconds_to_time)
            result_df['EasyTest秒數'] = result_df['EasyTest總時數'].apply(time_to_seconds)
            result_df.to_excel(result_path, index=False)

        else:
            return "不支援的檔案格式", 400

        @after_this_request
        def remove_files(response):
            for path in [history_path, online_info_path, result_path]:
                if os.path.exists(path):
                    os.remove(path)
            return response

        return send_file(result_path, as_attachment=True)

    except Exception as e:
        return f"處理檔案時發生錯誤: {str(e)}", 500

if __name__ == '__main__':
    cleanup_thread = Thread(target=cleanup_task, daemon=True)
    cleanup_thread.start()
    app.run(host='0.0.0.0', port=5000, debug=True)