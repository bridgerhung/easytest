import os
import time
from threading import Thread
from flask import Flask, request, send_file, render_template, after_this_request
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'

# 創建目錄
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

def delete_old_files(folder, age_in_seconds):
    """
    刪除資料夾中超過 age_in_seconds 的文件。
    """
    now = time.time()
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        if os.path.isfile(file_path):
            file_age = now - os.path.getmtime(file_path)
            if file_age > age_in_seconds:
                os.remove(file_path)
                print(f"Deleted: {file_path}")

def cleanup_task():
    """
    後台清理任務，每 1 小時執行一次。
    """
    while True:
        delete_old_files(UPLOAD_FOLDER, 60)  # 刪除超過 1 小時的文件
        delete_old_files(RESULT_FOLDER, 60)
        time.sleep(60)  # 每小時執行一次

@app.route('/legacy')
def index():
    return render_template('legacy.html')

@app.route('/legacy/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return {"error": "No file uploaded"}, 400

    file = request.files['file']
    if file.filename == '':
        return {"error": "No file selected"}, 400

    filename = secure_filename(file.filename)

    if filename.lower().endswith('.xlsx'):
        input_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(input_path)
        name, ext = os.path.splitext(filename)
        output_filename = f"{name}-count{ext}"
        output_path = os.path.join(RESULT_FOLDER, output_filename)
        wb = load_workbook(input_path)
        ws = wb.active

        # 秒數欄位 E
        ws["E2"] = "秒數"
        max_row = ws.max_row
        for i in range(3, max_row + 1):
            c_cell = f"C{i}"
            e_cell = f"E{i}"
            formula_e = '=IFERROR(LEFT({},FIND("天",{})-1)*86400,0) + IFERROR(MID({},FIND("天",{})+2,FIND("小時",{})-FIND("天",{})-2)*3600,0) + IFERROR(MID({},FIND("小時",{})+2,FIND("分",{})-FIND("小時",{})-2)*60,0) + IFERROR(MID({},FIND("分",{})+2,FIND("秒",{})-FIND("分",{})-2),0)'.format(
                c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell,c_cell
            )
            ws[e_cell] = formula_e

        # 分數欄位 F
        ws["F2"] = "分數"
        for i in range(3, max_row + 1):
            e_cell = f"E{i}"
            f_cell = f"F{i}"
            formula_f = '=IF({}>=43200, 5, ROUND({}/43200*5, 2))'.format(e_cell, e_cell)
            ws[f_cell] = formula_f

        # 時分欄位 G
        ws["G2"] = "時分"
        for i in range(3, max_row + 1):
            e_cell = f"E{i}"
            g_cell = f"G{i}"
            formula_g = '=TEXT(INT({}/3600),"00")&"時"&TEXT(INT(MOD({},3600)/60),"00")&"分"'.format(e_cell, e_cell)
            ws[g_cell] = formula_g

        wb.save(output_path)
        
        @after_this_request
        def remove_file(response):
            try:
                os.remove(output_path)
                os.remove(input_path)
            except Exception as e:
                print(f"Error deleting files: {e}")
            return response
        
        return send_file(output_path, as_attachment=True)
        
    else:
        # 處理 CSV 檔
        input_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(input_path)
        name, ext = os.path.splitext(filename)
        output_filename = f"{name}-count.xlsx"
        output_path = os.path.join(RESULT_FOLDER, output_filename)

        try:
            # 嘗試讀取 CSV 文件
            try:
                df = pd.read_csv(input_path, encoding='utf-8')  # 預設嘗試 UTF-8
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(input_path, encoding='big5')  # 嘗試 Big5
                except UnicodeDecodeError:
                    df = pd.read_csv(input_path, encoding='ISO-8859-1')  # 最後嘗試 ISO-8859-1
            # 設置 "總時數" 為文字
            if '總時數' in df.columns:
                df['總時數'] = df['總時數'].astype(str)

            # 刪除 "登入次數"
            if '登入次數' in df.columns:
                df = df.drop(columns=['登入次數'])

            # 將結果儲存為 XLSX
            df.to_excel(output_path, index=False)

            # 加載工作簿並添加公式
            wb = load_workbook(output_path)
            ws = wb.active

            # 秒數欄位
            ws["D1"] = "秒數"
            for i in range(2, len(df) + 2):
                c_cell = f"C{i}"
                d_cell = f"D{i}"
                formula_d = '=IFERROR(LEFT({},FIND("時",{})-1)*3600,0) + IFERROR(MID({},FIND("時",{})+1,FIND("分",{})-FIND("時",{})-1)*60,0)'.format(
                    c_cell, c_cell, c_cell, c_cell, c_cell, c_cell
                )
                ws[d_cell] = formula_d

            # 分數欄位
            ws["E1"] = "分數"
            for i in range(2, len(df) + 2):
                d_cell = f"D{i}"
                e_cell = f"E{i}"
                formula_e = '=IF({}>=43200, 5, ROUND({}/43200*5, 2))'.format(d_cell, d_cell)
                ws[e_cell] = formula_e

            wb.save(output_path)

        except Exception as e:
            return {"error": str(e)}, 500

        @after_this_request
        def remove_file(response):
            try:
                os.remove(output_path)
                os.remove(input_path)
            except Exception as e:
                print(f"Error deleting files: {e}")
            return response
        
        return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    # 啟動清理任務
    cleanup_thread = Thread(target=cleanup_task, daemon=True)
    cleanup_thread.start()
    app.run(host='0.0.0.0', port=6500, debug=True)
